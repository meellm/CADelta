"""Round-trip test for :func:`cadelta.gui.excel_report.write_excel_report`.

We build a known DiffResult, write the xlsx, then re-read it via openpyxl
and assert the summary counts and per-part rows match. The point isn't to
test openpyxl itself — it's to pin the contract that future refactors of
the report writer don't silently shift columns or drop rows.
"""
from __future__ import annotations

from pathlib import Path

import pytest

from cadelta.gui.excel_report import write_excel_report
from cadelta.matcher import Status, diff_parts
from cadelta.reader import load_parts, load_parts_with_doc

from .conftest import box_at, make_step


@pytest.fixture
def step_pair_with_diff(tmp_path: Path):
    """Build v1.step + v2.step with one of each status, run the diff, return
    everything the report writer needs."""
    v1 = tmp_path / "v1.step"
    v2 = tmp_path / "v2.step"
    make_step(v1, [
        ("A", box_at(10, 10, 10, 0, 0, 0)),
        ("B", box_at(15, 15, 15, 40, 0, 0)),
        ("C", box_at(20, 20, 20, 80, 0, 0)),
    ])
    make_step(v2, [
        ("A", box_at(10, 10, 10, 0, 0, 0)),
        ("B", box_at(15, 15, 15, 40.5, 0, 0)),  # moved
        ("D", box_at(12, 12, 12, 120, 0, 0)),   # added
    ])
    parts_v1 = load_parts(v1)
    parts_v2, _doc = load_parts_with_doc(v2)
    result = diff_parts(parts_v1, parts_v2)
    return v1, v2, result


def test_excel_report_roundtrip(step_pair_with_diff, tmp_path: Path):
    """Counts on the Summary sheet and row count on the Parts sheet must
    match the DiffResult. Spot-check a couple of cells to catch column-order
    regressions."""
    from openpyxl import load_workbook

    v1, v2, result = step_pair_with_diff
    out = tmp_path / "diff_report.xlsx"
    write_excel_report(result, out, v1_path=v1, v2_path=v2,
                       tol_mm=0.01, tol_deg=0.01)
    assert out.exists() and out.stat().st_size > 0

    wb = load_workbook(out)
    assert wb.sheetnames == ["Summary", "Parts"]

    ws_summary = wb["Summary"]
    # Find the count rows by their label column (rows 10–13 per current layout).
    summary_pairs = {
        ws_summary.cell(row=r, column=1).value: ws_summary.cell(row=r, column=2).value
        for r in range(10, 14)
    }
    expected_counts = {s: len(result.by_status(s)) for s in Status}
    assert summary_pairs["UNCHANGED"] == expected_counts[Status.UNCHANGED]
    assert summary_pairs["MOVED"] == expected_counts[Status.MOVED]
    assert summary_pairs["ADDED"] == expected_counts[Status.ADDED]
    assert summary_pairs["REMOVED"] == expected_counts[Status.REMOVED]

    ws_parts = wb["Parts"]
    # Header row + one data row per DiffEntry.
    assert ws_parts.max_row == 1 + len(result.entries), (
        f"Parts sheet should have {1 + len(result.entries)} rows; "
        f"got {ws_parts.max_row}"
    )
    # Header is in the documented order — protects column readers downstream.
    header = [ws_parts.cell(row=1, column=c).value for c in range(1, 11)]
    assert header == [
        "Status", "Name", "Δ mm", "Δ deg",
        "v1 X", "v1 Y", "v1 Z",
        "v2 X", "v2 Y", "v2 Z",
    ]

    # The MOVED entry's Δ mm column should match the diff's reported delta.
    statuses_in_sheet = [ws_parts.cell(row=r, column=1).value for r in range(2, ws_parts.max_row + 1)]
    moved_row = 2 + statuses_in_sheet.index("moved")
    moved_entry = result.by_status(Status.MOVED)[0]
    sheet_delta = ws_parts.cell(row=moved_row, column=3).value
    assert sheet_delta == pytest.approx(moved_entry.delta_mm, abs=1e-6)

    # The Status cell of every row carries a fill — cheap proof the
    # color-fill code path ran. We don't assert specific RGB bytes here
    # (those are pinned by the writer-flag tests via COLOR_BY_STATUS).
    for r in range(2, ws_parts.max_row + 1):
        fill = ws_parts.cell(row=r, column=1).fill
        assert fill.fill_type == "solid", (
            f"row {r} Status cell should have a solid fill; got {fill.fill_type}"
        )
