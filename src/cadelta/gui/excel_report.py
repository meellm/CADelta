"""Build a per-part Excel report from a :class:`cadelta.matcher.DiffResult`.

The CLI already writes a JSON report; this module covers the GUI's
"Excel report" toggle. The two reports carry the same data — the choice is
just about consumption: spreadsheet UI versus machine-readable.

Workbook layout:

- **Summary** sheet: counts (UNCHANGED / MOVED / ADDED / REMOVED), v1 and v2
  paths, tolerances used, generation timestamp.
- **Parts** sheet: one row per ``DiffEntry``. Columns:
  Status, Name, Δmm (translation), Δdeg (rotation), v1 X/Y/Z, v2 X/Y/Z.
  The Status cell is fill-colored to match the diff color used in
  ``diff.step`` so the spreadsheet reads the same way as the CAD output.

Pure Python — no native deps beyond ``openpyxl``. Lives in the GUI subpackage
because the CLI doesn't expose this output today; promote later if needed.
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from cadelta.matcher import DiffEntry, DiffResult, Status
from cadelta.writer import COLOR_BY_STATUS, COLOR_MOVED_FROM


# Header row for the Parts sheet. Kept as a module constant so the test can
# pin column order without parsing the writer code.
_PARTS_HEADER = [
    "Status",
    "Name",
    "Δ mm",
    "Δ deg",
    "v1 X", "v1 Y", "v1 Z",
    "v2 X", "v2 Y", "v2 Z",
]


def _rgb_to_hex(rgb: tuple[float, float, float]) -> str:
    """Convert an (r, g, b) float triple in [0, 1] to an openpyxl ARGB hex
    string (e.g. ``"FFFFD800"`` for yellow). The leading ``FF`` is the alpha
    channel — fully opaque."""
    r, g, b = (max(0, min(255, round(c * 255))) for c in rgb)
    return f"FF{r:02X}{g:02X}{b:02X}"


def _entry_color(entry: DiffEntry) -> tuple[float, float, float] | None:
    """Resolve the diff color a part would render with in ``diff.step``.

    Mirrors the writer's logic: each Status maps to one COLOR_BY_STATUS
    entry, with the special case that an entry whose v2 part carries an
    explicit color (and is unchanged) keeps that color. Returns ``None``
    when the entry has no color we'd render — currently never hits in
    practice, but keep the option open for callers that want to leave
    the cell unfilled.
    """
    if entry.status == Status.UNCHANGED and entry.part_v2 is not None and entry.part_v2.color is not None:
        return entry.part_v2.color
    return COLOR_BY_STATUS.get(entry.status)


def _entry_name(entry: DiffEntry) -> str:
    if entry.part_v2 is not None and entry.part_v2.name:
        return entry.part_v2.name
    if entry.part_v1 is not None and entry.part_v1.name:
        return entry.part_v1.name
    return ""


def _entry_centroid(part) -> tuple[float, float, float]:
    """Return a (x, y, z) tuple for a Part's world-space centroid, or
    ``("", "", "")`` if the part is missing. The empty-string fallback
    keeps the cell readable (Excel renders ``None`` as the literal text
    ``"None"`` which is uglier than a blank)."""
    if part is None or part.centroid is None:
        return ("", "", "")
    return (float(part.centroid[0]), float(part.centroid[1]), float(part.centroid[2]))


def write_excel_report(
    result: DiffResult,
    out_xlsx: Path,
    *,
    v1_path: Path,
    v2_path: Path,
    tol_mm: float,
    tol_deg: float,
) -> None:
    """Write an .xlsx report for ``result`` to ``out_xlsx``.

    The caller is responsible for ensuring ``out_xlsx``'s parent directory
    exists; we mkdir defensively below regardless.
    """
    # Local import keeps the dep optional for non-GUI callers (e.g. when
    # running the matcher tests, openpyxl shouldn't be required).
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    # --- Summary sheet ------------------------------------------------------
    ws_summary = wb.active
    ws_summary.title = "Summary"

    counts = {s: len(result.by_status(s)) for s in Status}
    summary_rows: Iterable[tuple] = [
        ("CADelta diff report", ""),
        ("Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("", ""),
        ("v1 file", str(v1_path)),
        ("v2 file", str(v2_path)),
        ("Translation tolerance (mm)", tol_mm),
        ("Rotation tolerance (deg)", tol_deg),
        ("", ""),
        ("Status", "Count"),
        ("UNCHANGED", counts[Status.UNCHANGED]),
        ("MOVED",     counts[Status.MOVED]),
        ("ADDED",     counts[Status.ADDED]),
        ("REMOVED",   counts[Status.REMOVED]),
    ]
    for row in summary_rows:
        ws_summary.append(row)
    # Bold the title and the inner header row.
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws_summary.cell(row=9, column=1).font = Font(bold=True)
    ws_summary.cell(row=9, column=2).font = Font(bold=True)
    # Reasonable width so paths and labels are legible without manual sizing.
    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 60

    # --- Parts sheet --------------------------------------------------------
    ws_parts = wb.create_sheet("Parts")
    ws_parts.append(_PARTS_HEADER)
    for col_idx in range(1, len(_PARTS_HEADER) + 1):
        ws_parts.cell(row=1, column=col_idx).font = Font(bold=True)
    # Freeze the header so it stays visible while scrolling — small QoL.
    ws_parts.freeze_panes = "A2"

    for entry in result.entries:
        v1_xyz = _entry_centroid(entry.part_v1)
        v2_xyz = _entry_centroid(entry.part_v2)
        ws_parts.append([
            entry.status.value,
            _entry_name(entry),
            entry.delta_mm,
            entry.delta_deg,
            v1_xyz[0], v1_xyz[1], v1_xyz[2],
            v2_xyz[0], v2_xyz[1], v2_xyz[2],
        ])
        # Fill the Status cell with the diff color so the spreadsheet
        # reads the same as diff.step. We deliberately fill only the
        # status column — filling the whole row would clash with the
        # user's own conditional formatting if they layer on top.
        rgb = _entry_color(entry)
        if rgb is not None:
            row_idx = ws_parts.max_row
            cell = ws_parts.cell(row=row_idx, column=1)
            argb = _rgb_to_hex(rgb)
            cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type="solid")

    # Modest column widths for readability — Excel's auto-width is opt-in
    # and runtime-only, so we set static widths.
    widths = [12, 40, 10, 10, 10, 10, 10, 10, 10, 10]
    for letter, width in zip("ABCDEFGHIJ", widths):
        ws_parts.column_dimensions[letter].width = width

    out_xlsx = Path(out_xlsx)
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
